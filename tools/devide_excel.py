import openpyxl

# 拆分所有的合并单元格，并赋予合并之前的值。
# 由于openpyxl并没有提供拆分并填充的方法，所以使用该方法进行完成
def unmerge_and_fill_cells(worksheet):
    all_merged_cell_ranges = list(
        worksheet.merged_cells.ranges
    )

    for merged_cell_range in all_merged_cell_ranges:
        merged_cell = merged_cell_range.start_cell
        worksheet.unmerge_cells(range_string=merged_cell_range.coord)

        for row_index, col_index in merged_cell_range.cells:
            cell = worksheet.cell(row=row_index, column=col_index)
            cell.value = merged_cell.value

# 读取原始xlsx文件，拆分并填充单元格，然后生成中间临时文件。
def unmerge_cell(filename):
    wb = openpyxl.load_workbook(filename)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        unmerge_and_fill_cells(sheet)
    filename = filename.replace(".xlsx", "_temp.xlsx")
    wb.save(filename)
    wb.close()

    return filename

if __name__ == '__main__':
    unmerge_cell("./Data.xlsx")